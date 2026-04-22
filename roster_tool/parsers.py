from __future__ import annotations

import hashlib
import re
from zipfile import BadZipFile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


TIMEZONE = ZoneInfo("Australia/Melbourne")

MMC_LOCATION = "MMC Car Park, Tarella Road, Clayton VIC 3168, Australia"
DDH_LOCATION = "DDH Car Park, 135 David St, Dandenong VIC 3175, Australia"

WEEKDAY_PREFIXES = ("Mon.", "Tue.", "Wed.", "Thu.", "Fri.", "Sat.", "Sun.")
TIME_PREFIX_RE = re.compile(r"^\s*(\d{2})(\d{2})-(\d{2})(\d{2})\s+(.+?)\s*$")

MMC_TEAM_MAP = {
    "G": "Green",
    "A": "Amber",
    "R": "Resus",
    "C": "Clinic",
}

DDH_LABEL_MAP = {
    "Clinical Support": "CS",
    "SSU SMS": "SSU",
    "Orange PM (on-call)": "Orange PM",
    "AVAO PM": "AVAO PM",
    "AVAO AM": "AVAO AM",
    "PM FAST IC": "FAST PM",
    "Orange AM IC": "Orange AM",
    "Silver AM IC": "Silver AM",
    "onsite CS": "CS onsite",
    "PHNW clinical": "PHNW",
    "PHNW": "PHNW",
}

WEEKLY_LEAVE_LABELS = {"ANNUAL LEAVE", "CONFERENCE LEAVE"}
IGNORED_EXACT = {
    "",
    "AL",
    "A/L",
    "EXAM",
    "EXAM LEAVE",
    "CME LEAVE",
    "PARENTAL LEAVE",
    "N/A",
}
IGNORED_CONTAINS = (
    "TEACHING",
    "EXAM",
    "MEETING",
    "MOCK",
    "HOLIDAY",
    "LABOUR DAY",
    "GOOD FRIDAY",
    "EASTER",
    "ANZAC",
    "BENDIGO",
    "ACEM",
    "REZA",
    "JENNY",
    "ARRHCHIE",
)
MMC_CROSS_SITE_PREFIX = "DANDENONG"
MMC_SECTION_BREAKS = {"GERIATRICIAN", "CMO", "SENIOR REG", "INTERMEDIATE REG", "JUNIOR REG", "HMO"}


@dataclass(frozen=True)
class DoctorOption:
    key: str
    display_name: str


@dataclass(frozen=True)
class CalendarEvent:
    source: str
    title: str
    start: datetime | date
    end: datetime | date
    all_day: bool
    location: str | None = None


def normalize_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", " ", value).strip().upper()
    return re.sub(r"\s+", " ", cleaned)


def detect_source_type(path: str | Path) -> str:
    try:
        workbook = load_workbook(path, data_only=True)
    except (InvalidFileException, BadZipFile, OSError) as exc:
        raise ValueError(
            f"{Path(path).name} is not a supported MMC workbook or Dandenong Hospital FindMyShift export."
        ) from exc
    sheetnames = set(workbook.sheetnames)
    if "Whole thing" in sheetnames and any(name.startswith("Week ") for name in sheetnames):
        return "mmc"
    sheet = workbook.active
    if sheet.max_column >= 8 and any(_is_date_row(sheet, idx) for idx in range(1, min(sheet.max_row, 4) + 1)):
        return "ddh"
    raise ValueError(
        f"{Path(path).name} is not a supported MMC workbook or Dandenong Hospital FindMyShift export."
    )


def doctor_options(
    mmc_path: str | Path | None = None,
    ddh_path: str | Path | None = None,
) -> list[DoctorOption]:
    if not mmc_path and not ddh_path:
        raise ValueError("Upload at least one MMC or DDH roster.")
    mmc_names = extract_mmc_names(mmc_path) if mmc_path else {}
    ddh_names = extract_ddh_names(ddh_path) if ddh_path else {}
    if mmc_path and ddh_path:
        common_keys = sorted(set(mmc_names) & set(ddh_names))
    elif mmc_path:
        common_keys = sorted(mmc_names)
    else:
        common_keys = sorted(ddh_names)
    options = []
    for key in common_keys:
        display = mmc_names.get(key) or ddh_names[key]
        options.append(DoctorOption(key=key, display_name=display))
    return options


def generate_events(
    mmc_path: str | Path | None,
    ddh_path: str | Path | None,
    doctor_key: str,
) -> list[CalendarEvent]:
    events: list[CalendarEvent] = []
    if mmc_path:
        events.extend(parse_mmc_events(mmc_path, doctor_key))
    if ddh_path:
        events.extend(parse_ddh_events(ddh_path, doctor_key))
    return sorted(events, key=_event_sort_key)


def export_ics(events: Iterable[CalendarEvent], doctor_display_name: str) -> str:
    now = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Roster to Calendar Tool//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:Roster - " + _escape_ics_text(doctor_display_name),
        "X-WR-TIMEZONE:Australia/Melbourne",
    ]
    for event in events:
        uid_seed = f"{event.source}|{event.title}|{event.start}|{event.end}|{event.location or ''}"
        uid = hashlib.sha1(uid_seed.encode("utf-8")).hexdigest() + "@roster-to-calendar"
        lines.extend(
            [
                "BEGIN:VEVENT",
                f"UID:{uid}",
                f"DTSTAMP:{now}",
                f"SUMMARY:{_escape_ics_text(event.title)}",
            ]
        )
        if event.all_day:
            start_date = _as_date(event.start)
            end_date = _as_date(event.end)
            lines.append(f"DTSTART;VALUE=DATE:{start_date.strftime('%Y%m%d')}")
            lines.append(f"DTEND;VALUE=DATE:{end_date.strftime('%Y%m%d')}")
        else:
            assert isinstance(event.start, datetime)
            assert isinstance(event.end, datetime)
            lines.append(
                "DTSTART;TZID=Australia/Melbourne:" + event.start.strftime("%Y%m%dT%H%M%S")
            )
            lines.append(
                "DTEND;TZID=Australia/Melbourne:" + event.end.strftime("%Y%m%dT%H%M%S")
            )
        if event.location:
            lines.append(f"LOCATION:{_escape_ics_text(event.location)}")
        lines.append("END:VEVENT")
    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"


def preview_summary(events: list[CalendarEvent]) -> dict[str, str | int]:
    if not events:
        return {"count": 0, "date_range": "No events found"}
    first = _as_date(events[0].start)
    last_event_end = _as_date(events[-1].end)
    last = last_event_end - timedelta(days=1) if events[-1].all_day else last_event_end
    return {
        "count": len(events),
        "date_range": f"{first.isoformat()} to {last.isoformat()}",
    }


def extract_mmc_names(path: str | Path | None) -> dict[str, str]:
    if not path:
        return {}
    workbook = load_workbook(path, data_only=True)
    names: dict[str, str] = {}
    for sheet in workbook.worksheets:
        if not sheet.title.startswith("Week "):
            continue
        for row_idx, raw_name in _iter_mmc_consultant_names(sheet):
            if not isinstance(raw_name, str):
                continue
            if not _looks_like_person_name(raw_name):
                continue
            key = normalize_name(raw_name)
            names.setdefault(key, raw_name.strip())
    return names


def extract_ddh_names(path: str | Path | None) -> dict[str, str]:
    if not path:
        return {}
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    names: dict[str, str] = {}
    for row_idx in range(1, sheet.max_row + 1):
        value = sheet.cell(row_idx, 1).value
        if not isinstance(value, str):
            continue
        value = value.strip()
        if not value or _is_date_row(sheet, row_idx):
            continue
        if not _looks_like_person_name(value):
            continue
        key = normalize_name(value)
        names.setdefault(key, value)
    return names


def parse_mmc_events(path: str | Path, doctor_key: str) -> list[CalendarEvent]:
    workbook = load_workbook(path, data_only=True)
    events: list[CalendarEvent] = []
    for sheet in workbook.worksheets:
        if not sheet.title.startswith("Week "):
            continue
        week_dates = [_coerce_date(sheet.cell(4, col).value) for col in range(6, 13)]
        if any(day is None for day in week_dates):
            continue
        week_dates = [day for day in week_dates if day is not None]
        for row_idx, raw_name in _iter_mmc_consultant_names(sheet):
            if not isinstance(raw_name, str):
                continue
            if normalize_name(raw_name) != doctor_key:
                continue
            raw_week = [_clean_text(sheet.cell(row_idx, col).value) for col in range(6, 13)]
            if not any(raw_week):
                continue
            weekly_leave = _first_weekly_leave(raw_week)
            if weekly_leave:
                monday = week_dates[0]
                events.append(
                    CalendarEvent(
                        source="MMC",
                        title=f"MMC: {weekly_leave.title()}",
                        start=monday,
                        end=monday + timedelta(days=7),
                        all_day=True,
                    )
                )
                continue
            for day, raw in zip(week_dates, raw_week):
                if not raw:
                    continue
                event = _parse_mmc_entry(day, raw)
                if event:
                    events.append(event)
            break
    return events


def assign_sources(paths: Iterable[str | Path]) -> dict[str, str | None]:
    paths = list(paths)
    if len(paths) > 2:
        raise ValueError("Upload no more than two roster files at a time.")
    assignments: dict[str, str | None] = {"mmc": None, "ddh": None}
    for path in paths:
        source = detect_source_type(path)
        if assignments[source] is not None:
            raise ValueError(f"Upload at most one {source.upper()} roster at a time.")
        assignments[source] = str(path)
    if assignments["mmc"] is None and assignments["ddh"] is None:
        raise ValueError("Upload at least one MMC or DDH roster.")
    return assignments


def parse_ddh_events(path: str | Path, doctor_key: str) -> list[CalendarEvent]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    events: list[CalendarEvent] = []
    row_idx = 1
    while row_idx <= sheet.max_row:
        if not _is_date_row(sheet, row_idx):
            row_idx += 1
            continue
        week_dates = [_parse_ddh_date(sheet.cell(row_idx, col).value) for col in range(2, 9)]
        if any(day is None for day in week_dates):
            row_idx += 1
            continue
        name_row = row_idx + 1
        if name_row > sheet.max_row:
            break
        raw_name = sheet.cell(name_row, 1).value
        label_row = [_clean_text(sheet.cell(name_row, col).value) for col in range(2, 9)]
        time_row_values = [_clean_text(sheet.cell(name_row + 1, col).value) for col in range(2, 9)] if name_row + 1 <= sheet.max_row else [""] * 7
        has_time_row = name_row + 1 <= sheet.max_row and not _is_date_row(sheet, name_row + 1) and any(time_row_values)
        if isinstance(raw_name, str) and normalize_name(raw_name) == doctor_key:
            weekly_leave = _first_weekly_leave(label_row)
            if weekly_leave:
                monday = week_dates[0]
                events.append(
                    CalendarEvent(
                        source="DDH",
                        title=f"DDH: {weekly_leave.title()}",
                        start=monday,
                        end=monday + timedelta(days=7),
                        all_day=True,
                    )
                )
            else:
                time_row = time_row_values if has_time_row else [""] * 7
                for day, label, time_text in zip(week_dates, label_row, time_row):
                    if not label and not time_text:
                        continue
                    event = _parse_ddh_entry(day, label, time_text)
                    if event:
                        events.append(event)
        row_idx = row_idx + (3 if has_time_row else 2)
    return events


def _parse_mmc_entry(day: date, raw: str) -> CalendarEvent | None:
    upper = raw.upper()
    if _should_ignore_mmc(raw):
        return None
    if upper == "PHNW":
        return CalendarEvent("MMC", "MMC: PHNW", day, day + timedelta(days=1), True)
    if upper.startswith("S/L"):
        return CalendarEvent("MMC", f"MMC: {_normalize_sick_leave_label(raw)}", day, day + timedelta(days=1), True)

    explicit = _extract_time_prefix(raw)
    label = explicit["label"] if explicit else raw.strip()
    normalized = _normalize_mmc_label(label)
    if not normalized:
        return None

    title = f"MMC: {normalized['title']}"
    location = normalized["location"]

    if explicit:
        start_dt, end_dt = _build_datetimes(day, explicit["start"], explicit["end"])
        return CalendarEvent("MMC", title, start_dt, end_dt, False, location)

    if normalized["all_day"]:
        return CalendarEvent("MMC", title, day, day + timedelta(days=1), True, location)

    start_time, end_time = normalized["default_times"]
    start_dt, end_dt = _build_datetimes(day, start_time, end_time)
    return CalendarEvent("MMC", title, start_dt, end_dt, False, location)


def _parse_ddh_entry(day: date, label: str, time_text: str) -> CalendarEvent | None:
    if not label:
        return None
    if _parse_ddh_time_row(label):
        return _parse_ddh_entry(day, time_text, label) if time_text else None
    upper = label.upper()
    if upper in {"AM", "PM"}:
        return None
    if upper == "PHNW" or upper == "PHNW CLINICAL":
        return CalendarEvent("DDH", "DDH: PHNW", day, day + timedelta(days=1), True)
    if upper.startswith("S/L"):
        return CalendarEvent("DDH", f"DDH: {_normalize_sick_leave_label(label)}", day, day + timedelta(days=1), True)
    if _should_ignore_common(label):
        return None

    mapped = DDH_LABEL_MAP.get(label, label)
    title = f"DDH: {mapped}"
    location = _ddh_location_for_title(mapped)

    if time_text:
        parsed = _parse_ddh_time_row(time_text)
        if parsed:
            start_dt, end_dt = _build_datetimes(day, parsed[0], parsed[1])
            return CalendarEvent("DDH", title, start_dt, end_dt, False, location)
    return CalendarEvent("DDH", title, day, day + timedelta(days=1), True, location)


def _normalize_mmc_label(label: str) -> dict[str, object] | None:
    code = label.strip().upper()
    if code == "CS":
        return {"title": "CS", "location": None, "all_day": True, "default_times": None}
    if code == "CSO":
        return {"title": "CSO", "location": MMC_LOCATION, "all_day": True, "default_times": None}
    if len(code) == 4 and code[1:3] == "SS" and code[0] in {"A", "P"} and code[3] in {"C", "R"}:
        shift_label = "AM" if code[0] == "A" else "PM"
        float_suffix = " Float" if code[3] == "R" else ""
        default_times = ((7, 30), (17, 30)) if code[0] == "A" else ((14, 30), (0, 0))
        return {
            "title": f"SSU {shift_label}{float_suffix}",
            "location": MMC_LOCATION,
            "all_day": False,
            "default_times": default_times,
        }
    if len(code) == 3 and code[0] in {"A", "P"} and code[1] in MMC_TEAM_MAP and code[2] in {"C", "R"}:
        shift_label = "AM" if code[0] == "A" else "PM"
        float_suffix = " Float" if code[2] == "R" else ""
        default_times = ((8, 0), (17, 30)) if code[0] == "A" else ((14, 30), (0, 0))
        return {
            "title": f"{MMC_TEAM_MAP[code[1]]} {shift_label}{float_suffix}",
            "location": MMC_LOCATION,
            "all_day": False,
            "default_times": default_times,
        }
    return None


def _ddh_location_for_title(title: str) -> str | None:
    if title in {"CS", "PHNW"}:
        return None
    if title == "CS onsite":
        return DDH_LOCATION
    return DDH_LOCATION


def _parse_ddh_time_row(value: str) -> tuple[tuple[int, int], tuple[int, int]] | None:
    match = re.match(r"^\s*(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})\s*$", value)
    if not match:
        return None
    return (int(match.group(1)), int(match.group(2))), (int(match.group(3)), int(match.group(4)))


def _parse_ddh_date(value: object) -> date | None:
    if not isinstance(value, str):
        return None
    value = value.replace("Sept.", "Sep.").replace("June.", "Jun.")
    for fmt in ("%a. %b. %d, %Y", "%a %b %d, %Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _is_date_row(sheet, row_idx: int) -> bool:
    first_value = sheet.cell(row_idx, 2).value
    return isinstance(first_value, str) and first_value.startswith(WEEKDAY_PREFIXES)


def _first_weekly_leave(values: Iterable[str]) -> str | None:
    for value in values:
        upper = value.upper()
        if upper in WEEKLY_LEAVE_LABELS:
            return value
    return None


def _coerce_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _extract_time_prefix(raw: str) -> dict[str, object] | None:
    match = TIME_PREFIX_RE.match(raw)
    if not match:
        return None
    start = (int(match.group(1)), int(match.group(2)))
    end = (int(match.group(3)), int(match.group(4)))
    return {"start": start, "end": end, "label": match.group(5).strip()}


def _build_datetimes(day: date, start_hm: tuple[int, int], end_hm: tuple[int, int]) -> tuple[datetime, datetime]:
    start_dt = datetime.combine(day, time(start_hm[0], start_hm[1]), tzinfo=TIMEZONE)
    end_day = day + timedelta(days=1) if end_hm <= start_hm else day
    end_dt = datetime.combine(end_day, time(end_hm[0], end_hm[1]), tzinfo=TIMEZONE)
    return start_dt, end_dt


def _should_ignore_mmc(raw: str) -> bool:
    upper = raw.strip().upper()
    if upper.startswith(MMC_CROSS_SITE_PREFIX):
        return True
    return _should_ignore_common(raw)


def _should_ignore_common(raw: str) -> bool:
    upper = raw.strip().upper()
    if upper in IGNORED_EXACT:
        return True
    return any(fragment in upper for fragment in IGNORED_CONTAINS)


def _looks_like_person_name(value: str) -> bool:
    cleaned = value.strip()
    if len(cleaned) < 5:
        return False
    if any(token in cleaned.upper() for token in ("NOT USED", "SMS", "DATE", "WEEK", "ROLE", "PAGER")):
        return False
    return bool(re.search(r"[A-Za-z]", cleaned) and " " in cleaned)


def _iter_mmc_consultant_names(sheet) -> Iterable[tuple[int, str]]:
    for row_idx in range(7, sheet.max_row + 1):
        section_marker = _clean_text(sheet.cell(row_idx, 3).value).upper()
        if section_marker in MMC_SECTION_BREAKS:
            break
        raw_name = sheet.cell(row_idx, 4).value
        if raw_name is None:
            continue
        yield row_idx, raw_name


def _event_sort_key(event: CalendarEvent) -> tuple[date, int, str]:
    start_date = _as_date(event.start)
    return start_date, 0 if event.all_day else 1, event.title


def _escape_ics_text(value: str) -> str:
    return (
        value.replace("\\", "\\\\")
        .replace(";", r"\;")
        .replace(",", r"\,")
        .replace("\n", r"\n")
    )


def _as_date(value: datetime | date) -> date:
    return value.date() if isinstance(value, datetime) else value


def _normalize_sick_leave_label(value: str) -> str:
    upper = value.strip().upper()
    suffix = upper.removeprefix("S/L").strip()
    return f"Sick Leave {suffix}".strip()
